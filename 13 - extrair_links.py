import openpyxl
import json
import os
from datetime import datetime

# Instruções de Execução:
# 1. Certifique-se de ter o Python instalado em seu sistema. Recomenda-se a versão 3.6 ou superior.
# 2. Instale as bibliotecas necessárias usando o seguinte comando:
#    pip install openpyxl
# 3. Altere o caminho do arquivo Excel (excel_file_path) para o local onde seu arquivo está armazenado.
# 4. Defina a variável extract_AllColumns para extrair todas as colunas, ou defina as colunas específicas na variável columns_to_extract.
# 5. Se search_text estiver vazio, o comportamento de filtragem será ignorado.
# 6. Execute o script. O arquivo JSON será gerado no diretório especificado.
# 7. Comando para executar: python '.\13 - extrair_links.py' ou python3 '.\13 - extrair_links.py'

# Nova variável com o texto a ser procurado
search_text = "https://butterflygrowth.sharepoint.com"  # Se estiver vazio, o comportamento de filtragem será ignorado

# Caminho do arquivo Excel
excel_file_path = r"C:\dev\scripts\ScriptsUteis\ArquivoExcel\Cronograma-Inbound-agger.xlsx"

# Carregar o arquivo Excel
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook["INFOCAP Atualização post"]

# Inicializar a lista para armazenar os dados
data = []

# Definir as colunas que você deseja extrair (por exemplo, ["B", "C"])
columns_to_extract = ["B", "C"]  # Defina aqui as colunas desejadas, se necessário

# Variável de controle para extrair todas as colunas ou somente as definidas
extract_AllColumns = False  # Defina como True para extrair todas as colunas

# Verificar quais colunas possuem ao menos uma célula com conteúdo
valid_columns = []

# Função para converter datetime para string
def convert_datetime(cell_value):
    if isinstance(cell_value, datetime):
        return cell_value.isoformat()  # Converte para o formato ISO 8601
    return cell_value

# Função para verificar se uma coluna tem conteúdo
def is_column_valid(sheet, col_letter):
    for row in sheet.iter_rows(min_row=2, min_col=openpyxl.utils.column_index_from_string(col_letter), 
                               max_col=openpyxl.utils.column_index_from_string(col_letter)):
        if any(cell.value is not None for cell in row):
            return True
    return False

# Iterar sobre as colunas e verificar se têm conteúdo
if extract_AllColumns:
    for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row):
        if any(cell.value is not None for cell in col):
            valid_columns.append(col[0].column_letter)

# Iterar sobre as linhas da planilha, começando da segunda linha
for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Ignorar a primeira linha (cabeçalho)
    row_data = {'linha': row_index}  # Adiciona o número da linha ao dicionário
    columns_to_process = valid_columns if extract_AllColumns else columns_to_extract

    # Verifica se todas as colunas especificadas em columns_to_extract são nulas
    all_columns_null = True

    for col in columns_to_process:
        col_index = openpyxl.utils.column_index_from_string(col) - 1  # Converter letra para índice
        cell = row[col_index]

        if cell.hyperlink:  # Verifica se a célula contém um hiperlink
            row_data[col] = cell.hyperlink.target  # Obtém o link real
            all_columns_null = False  # Pelo menos uma célula não é nula
        else:
            cell_value = convert_datetime(cell.value) if cell.value is not None else None
            row_data[col] = cell_value
            if cell_value is not None:
                all_columns_null = False  # Pelo menos uma célula não é nula

    # Se não estiver extraindo todas as colunas, só adicionar a linha se houver pelo menos uma célula não nula
    if not all_columns_null:
        data.append(row_data)

# Caminho do diretório para salvar o JSON
output_directory = r"C:\dev\scripts\ScriptsUteis\ArquivoExcel\files"
os.makedirs(output_directory, exist_ok=True)  # Cria o diretório se não existir

# Nome do arquivo JSON baseado na aba
json_file_name = f"{sheet.title}.json"
json_file_path = os.path.join(output_directory, json_file_name)

# Gravar os dados em um arquivo JSON
with open(json_file_path, "w", encoding="utf-8") as json_file:
    json.dump(data, json_file, ensure_ascii=False, indent=4)

print(f"Arquivo JSON gerado com sucesso em {json_file_path}!")

# Executar filtragem apenas se search_text não estiver vazio
if search_text:
    # Inicializar lista para armazenar os resultados encontrados
    filtered_data = [item for item in data if any(search_text in str(value) for value in item.values())]

    # Caminho e nome do novo arquivo JSON para os resultados filtrados
    filtered_json_file_name = f"{sheet.title}_filtered.json"
    filtered_json_file_path = os.path.join(output_directory, filtered_json_file_name)

    # Gravar os resultados filtrados em um novo arquivo JSON
    with open(filtered_json_file_path, "w", encoding="utf-8") as filtered_json_file:
        json.dump(filtered_data, filtered_json_file, ensure_ascii=False, indent=4)

    print(f"Arquivo JSON filtrado gerado com sucesso em {filtered_json_file_path}!")
else:
    print("search_text está vazio. O comportamento de filtragem foi ignorado.")
